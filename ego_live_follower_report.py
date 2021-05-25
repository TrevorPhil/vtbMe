# Author: 最上静香
# API impletemented by https://vtbs.moe/about.
# Version: v0.0.5

import requests as r
import pandas as pd
import os
import json
import time
from datetime import date
from openpyxl import load_workbook

def append_df_to_excel(filename, df, sheet_name, startcol=None,**to_excel_kwargs):
    # Excel file doesn't exist 
    # Or the excel file is empty
    # Save and exit
    if not os.path.isfile(filename) or pd.read_excel(filename).empty:
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startcol=startcol if startcol is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startcol is None and sheet_name in writer.book.sheetnames:
        startcol = writer.book[sheet_name].max_column
        
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startcol is None:
        startcol = 0
    else:
        df = df[df.columns[1]].to_frame()
        
    # erase title when start appending
    # write out the new sheet
    df.to_excel(writer, sheet_name, startcol=startcol, **to_excel_kwargs)

    # save the workbook
    writer.save()



vtb_full_info_url = "https://api.vtbs.moe/v1/info"
ego_room_ids = [7194103, 1086621, 22631364, 475577, 22588330, 11312, 10413051, 673595, 22572737, 22580086, 3923305, 22800243, 22595698, 3000303, 22605289, 52813, 22707677, 22805801, 22865894, 22620570, 22934732]

def decode_ego_info():
    with open('vtb_info.json', encoding='utf_8_sig') as file:
        vtbjson = file.read()

    decoded_ego_live_members =  json.loads(vtbjson, object_hook=ego_decoder)
    ego_live_info_list = list(filter(None, decoded_ego_live_members))
    return ego_live_info_list

def get_vtb_info():
    vtb = r.get(vtb_full_info_url).json()
    vtb_json_file = json.dumps(vtb)
    write_file_to_local(vtb_json_file)

def write_file_to_local(vtb):
    with open('vtb_info.json', 'w', encoding='utf_8_sig') as file:
        file.write(vtb)


# def delete_local_file():
#     if os.path.exists('vtb_info.txt'):
#         os.remove('vtb_info.txt')

def ego_decoder(dct):
    if 'roomid' in dct:
        if dct['roomid'] in ego_room_ids:
            return dct

def generate_daily_statistics(ego_live_info_list, **generate_kwargs) -> list:
    current_date_time = date.today().strftime('%Y.%m.%d')
    if 'change' in generate_kwargs:
        current_date_time_plus_description = current_date_time + '粉丝数变化'
    elif 'follower' in generate_kwargs:
        current_date_time_plus_description = current_date_time + '粉丝数'
    daily_change_list = list()
    for ego_members_info in ego_live_info_list:
        description = ego_members_info['follower'] if 'follower' in generate_kwargs else ego_members_info['rise']
        daily_change_list.append({'Name': ego_members_info['uname'], current_date_time_plus_description: description})
    return daily_change_list
    
def create_dataframe(df_list):
    df = pd.DataFrame(df_list)
    return df


get_vtb_info()
ego_list = decode_ego_info()

follower_list = generate_daily_statistics(ego_list, follower = True)
follower_change_list = generate_daily_statistics(ego_list, change = True)

follower_df = create_dataframe(follower_list)
follower_change_df = create_dataframe(follower_change_list)

append_df_to_excel('./ego_live_report.xlsx', follower_df, 'Sheet1', index = False)
append_df_to_excel('./ego_live_report.xlsx', follower_change_df, 'Sheet2', index = False)



